"""
QuickPic AutoSpec — Batch fetch specs via Excel export.
Navigates the brand > model > variant tree, selects up to 4, exports Excel.
Ultra-random human-like delays to stay safe.

Usage:
  python quickpic_fetch_specs.py              # Process all brands
  python quickpic_fetch_specs.py --brand vw   # Process single brand
  python quickpic_fetch_specs.py --limit 20   # Max variants to process
"""
import asyncio
import random
import os
import json
import re
import sys
import argparse
from datetime import datetime
from collections import defaultdict
import functools
print = functools.partial(print, flush=True)

try:
    import openpyxl
except ImportError:
    openpyxl = None  # Only needed for legacy Excel mode

from playwright.async_api import async_playwright

BASE_DIR = os.path.dirname(__file__)
CATALOGUE_FILE = os.path.join(BASE_DIR, "quickpic_catalogue.json")
RAW_DIR = os.path.join(BASE_DIR, "raw_data")
SPECS_DIR = os.path.join(BASE_DIR, "quickpic_specs")
EXCEL_DIR = os.path.join(BASE_DIR, "quickpic_excels")
PROGRESS_FILE = os.path.join(BASE_DIR, "quickpic_progress.json")
os.makedirs(SPECS_DIR, exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)

USERNAME = "dealerdigitalza@gmail.com"
PASSWORD = "LukeJason#1"


async def wild_delay(context=""):
    """Human-like delays between batches. Varies like real browsing."""
    r = random.random()
    if r < 0.25:
        d = random.uniform(3, 7)
    elif r < 0.55:
        d = random.uniform(8, 15)
    elif r < 0.80:
        d = random.uniform(16, 25)
    elif r < 0.93:
        d = random.uniform(26, 40)
    else:
        d = random.uniform(41, 65)
    label = f" ({context})" if context else ""
    print(f"    ... {d:.0f}s{label}")
    await asyncio.sleep(d)


async def medium_delay():
    """Medium delay for within-page actions."""
    d = random.uniform(2, 6)
    await asyncio.sleep(d)


async def small_delay():
    """Small delay for quick UI interactions."""
    d = random.uniform(0.5, 2)
    await asyncio.sleep(d)


async def human_type(locator, text):
    """Type with variable speed."""
    await locator.click()
    await asyncio.sleep(random.uniform(0.3, 1))
    base_speed = random.randint(40, 180)
    for ch in text:
        ms = base_speed + random.randint(-20, 60)
        if random.random() < 0.08:
            ms = random.randint(400, 900)
        await locator.press_sequentially(ch, delay=max(30, ms))
    await asyncio.sleep(random.uniform(0.5, 2))


async def dump_tree_items(page):
    """Debug: dump all visible tree items in the compare dialog."""
    items = await page.evaluate("""() => {
        const result = [];
        const modals = document.querySelectorAll('.w3-modal');
        for (const m of modals) {
            if (!m.innerText.includes('Compare Vehicles')) continue;
            // Check all items with various selectors
            const selectors = ['.chkboxItem', '.treeItem', 'li', '[class*=tree]', '[class*=node]', '[class*=item]'];
            const seen = new Set();
            for (const sel of selectors) {
                const els = m.querySelectorAll(sel);
                for (const el of els) {
                    const key = el.outerHTML.substring(0, 200);
                    if (seen.has(key)) continue;
                    seen.add(key);
                    const directText = Array.from(el.childNodes)
                        .filter(n => n.nodeType === 3)
                        .map(n => n.textContent.trim())
                        .filter(t => t)
                        .join(' ');
                    result.push({
                        selector: sel,
                        tag: el.tagName,
                        className: el.className.substring(0, 100),
                        directText: directText.substring(0, 80),
                        innerText: el.innerText.substring(0, 80).replace(/\\n/g, ' | '),
                        visible: el.offsetHeight > 0,
                        depth: el.closest('[class*=tree]')?.className?.substring(0, 50) || '',
                    });
                }
            }
            // Also get raw modal HTML structure (first 3000 chars)
            result.push({selector: 'MODAL_HTML', tag: 'modal', className: '', directText: '',
                         innerText: m.innerHTML.substring(0, 3000).replace(/\\n/g, ' '), visible: true, depth: ''});
        }
        return result;
    }""")
    return items


async def click_tree_item(page, text_to_find):
    """Click a brand or model item in the compare dialog tree.
    Items have class='chkboxItem' and contain the text like 'Volkswagen' or 'Polo (9)'.
    Returns True if clicked successfully."""
    result = await page.evaluate("""(textToFind) => {
        // Find the compare dialog (modal 2 with "Compare Vehicles")
        const modals = document.querySelectorAll('.w3-modal');
        for (const m of modals) {
            if (!m.innerText.includes('Compare Vehicles')) continue;
            const items = m.querySelectorAll('.chkboxItem');
            for (const item of items) {
                // Get the direct text (brand/model name without child counts)
                const directText = Array.from(item.childNodes)
                    .filter(n => n.nodeType === 3)
                    .map(n => n.textContent.trim())
                    .join('');
                if (directText === textToFind || item.textContent.trim().startsWith(textToFind + ' ') ||
                    item.textContent.trim().startsWith(textToFind + '(')) {
                    item.click();
                    return true;
                }
            }
        }
        return false;
    }""", text_to_find)
    return result


async def check_variant_checkbox(page, variant_name):
    """Check a variant's checkbox in the expanded model tree.
    Returns True if checked successfully."""
    result = await page.evaluate("""(variantName) => {
        const modals = document.querySelectorAll('.w3-modal');
        for (const m of modals) {
            if (!m.innerText.includes('Compare Vehicles')) continue;
            const checkboxes = m.querySelectorAll("input[type='checkbox']");
            for (const cb of checkboxes) {
                const parent = cb.closest('.chkboxItem') || cb.parentElement;
                if (!parent) continue;
                const text = parent.innerText.trim();
                if (text === variantName || text.startsWith(variantName + '\\n')) {
                    if (!cb.checked) {
                        cb.click();
                        return {checked: true, text: text.substring(0, 80)};
                    }
                    return {checked: true, alreadyChecked: true, text: text.substring(0, 80)};
                }
            }
        }
        return {checked: false};
    }""", variant_name)
    return result


async def click_show_button(page):
    """Click the Show button in the compare dialog."""
    result = await page.evaluate("""() => {
        const modals = document.querySelectorAll('.w3-modal');
        for (const m of modals) {
            if (!m.innerText.includes('Compare Vehicles')) continue;
            const buttons = m.querySelectorAll('button');
            for (const btn of buttons) {
                if (btn.textContent.trim() === 'Show') {
                    btn.click();
                    return true;
                }
            }
        }
        return false;
    }""")
    return result


async def scrape_comparison_html(page, selected_variants):
    """Scrape the comparison table rendered on page after clicking Show.
    selected_variants is a list of (brand_key, qp_brand, model_name, variant) tuples
    that were successfully selected (in order).
    Returns list of {variant: str, specs: {key: val}} dicts."""

    # The table is a simple 2+ column layout:
    # Column 0 = spec label, Columns 1-4 = values for each vehicle
    # Vehicle names are NOT in the table — they're in headers above
    data = await page.evaluate("""() => {
        const tables = document.querySelectorAll('table');
        if (tables.length === 0) return {rows: [], num_vehicles: 0};

        // Find the largest table
        let biggest = tables[0];
        for (const t of tables) {
            if (t.innerHTML.length > biggest.innerHTML.length) biggest = t;
        }

        // Extract cell text, stripping ranking button elements
        function getCellText(cell) {
            // Recursively extract text, skipping <button> elements (ranking badges)
            function extractText(el) {
                let parts = [];
                for (const node of el.childNodes) {
                    if (node.nodeType === 3) {
                        const t = node.textContent.trim();
                        if (t) parts.push(t);
                    } else if (node.nodeType === 1) {
                        const tag = node.tagName.toLowerCase();
                        // Skip button elements (these contain the 1-4 ranking numbers)
                        if (tag === 'button') continue;
                        // Skip score containers
                        const cls = (typeof node.className === 'string' ? node.className : '').toLowerCase();
                        if (cls.includes('score') || cls.includes('rank')) continue;
                        // Recurse into other elements
                        parts.push(extractText(node));
                    }
                }
                return parts.join(' ').trim();
            }
            return extractText(cell);
        }

        const rows = biggest.querySelectorAll('tr');
        const result = [];
        let num_vehicles = 0;

        for (const row of rows) {
            const cells = row.querySelectorAll('td, th');
            const texts = Array.from(cells).map(c => getCellText(c));
            if (texts.length >= 2) {
                result.push(texts);
                const nonEmpty = texts.slice(1).filter(t => t && t !== '-');
                if (nonEmpty.length > num_vehicles) num_vehicles = nonEmpty.length;
            }
        }

        return {rows: result, num_vehicles: num_vehicles};
    }""")

    rows = data.get("rows", [])
    num_vehicles = data.get("num_vehicles", 0)

    if not rows:
        print("    No table data found")
        return []

    if num_vehicles == 0:
        num_vehicles = len(selected_variants)

    print(f"    Table: {len(rows)} rows, {num_vehicles} vehicle columns")

    # Build spec dicts for each vehicle column
    vehicles = []
    for vi in range(num_vehicles):
        # Use the variant info from what we selected
        if vi < len(selected_variants):
            _, _, model_name, variant = selected_variants[vi]
            variant_name = variant["name"]
        else:
            variant_name = f"Vehicle {vi+1}"

        specs = {}
        for row in rows:
            if len(row) < 2:
                continue
            label = row[0].strip()
            if not label or len(label) < 2:
                continue
            # Skip section headers (General, Engine, etc.)
            if label in ("General", "Engine", "Safety", "Warranty", "Features",
                         "Aesthetics And Comfort", "Specifications", "Wheels And Tyres",
                         "Media Clips", "Score:"):
                continue

            col_idx = vi + 1
            if col_idx < len(row):
                val = row[col_idx].strip()
                if val and val not in ("-", "N/A", "n/a", "--", ""):
                    specs[label] = val

        vehicles.append({"variant": variant_name, "specs": specs})
        print(f"    Vehicle: {variant_name[:60]} — {len(specs)} specs")

    return vehicles


def slugify_model(name: str) -> str:
    """Convert model name to file slug."""
    s = name.lower().strip()
    s = re.sub(r'[^a-z0-9\-]+', '-', s)
    s = re.sub(r'-+', '-', s).strip('-')
    return s


XML_TO_SNAKE = {
    "EngineCapacity": "engine_capacity",
    "PowerKW": "power_kw",
    "TorqueNM": "torque_nm",
    "FuelType": "fuel_type",
    "Cylinders": "cylinders",
    "Transmission": "transmission",
    "Drivetrain": "drivetrain",
    "TopSpeedKMH": "top_speed_kmh",
    "Acceleration0to100": "acceleration_0_100",
    "FuelConsumptionL100KM": "fuel_consumption_l100km",
    "CO2EmissionsGKM": "co2_emissions_gkm",
    "LengthMM": "length_mm",
    "WidthMM": "width_mm",
    "HeightMM": "height_mm",
    "WheelbaseMM": "wheelbase_mm",
    "BootCapacityL": "boot_capacity_l",
    "KerbWeightKG": "kerb_weight_kg",
    "FuelTankL": "fuel_tank_l",
    "GroundClearanceMM": "ground_clearance_mm",
    "TurningCircleM": "turning_circle_m",
    "Airbags": "airbags",
    "ABS": "abs",
    "StabilityControl": "stability_control",
    "Warranty": "warranty",
    "ServicePlan": "service_plan",
}


def merge_specs_to_raw(brand_key, model_name, variant_name, spec_dict):
    """Merge scraped specs into the corresponding raw_data JSON file."""
    model_slug = slugify_model(model_name)
    raw_file = os.path.join(RAW_DIR, brand_key, f"{model_slug}.json")

    if not os.path.exists(raw_file):
        print(f"      No raw_data file for {brand_key}/{model_slug}")
        return False

    with open(raw_file) as f:
        raw = json.load(f)

    # Convert QuickPic specs to our feed format, then to snake_case keys
    feed_specs = specs_to_feed_format(brand_key, model_name,
                                      {"variant": variant_name, "specs": spec_dict})
    mapped_specs = {}
    for xml_key, val in feed_specs.get("specifications", {}).items():
        if val:
            snake = XML_TO_SNAKE.get(xml_key, xml_key.lower())
            mapped_specs[snake] = val

    matched = False
    for variant in raw.get("variants", []):
        if variant["variant_name"] == variant_name:
            variant["specs"] = mapped_specs
            matched = True
            break

    if not matched:
        # Fuzzy match: check if variant_name is contained in any variant
        for variant in raw.get("variants", []):
            if variant_name in variant["variant_name"] or variant["variant_name"] in variant_name:
                variant["specs"] = mapped_specs
                matched = True
                break

    if matched:
        with open(raw_file, "w") as f:
            json.dump(raw, f, indent=2, ensure_ascii=False)
        return True
    else:
        print(f"      No matching variant for '{variant_name}' in {raw_file}")
        return False


def parse_excel(filepath):
    """Parse AutoSpec Excel export into structured data."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    vehicles = []

    for col_idx in range(2, 6):  # Columns B(2) through E(5)
        brand_cell = ws.cell(row=3, column=col_idx).value
        variant_cell = ws.cell(row=4, column=col_idx).value
        if not brand_cell or not variant_cell:
            continue

        brand = str(brand_cell).strip()
        variant = str(variant_cell).strip()
        specs = {}

        for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
            label_cell = row[0]
            value_cell = row[col_idx - 1]

            if label_cell.value and value_cell.value:
                key = str(label_cell.value).strip()
                val = str(value_cell.value).strip()
                if val and val != "None":
                    specs[key] = val

        if specs:
            vehicles.append({
                "brand": brand,
                "variant": variant,
                "specs": specs,
            })

    return vehicles


def specs_to_feed_format(brand_key, model_name, variant_data):
    """Convert QuickPic spec dict to our feed's spec format."""
    s = variant_data.get("specs", {})

    def get(key, default=""):
        return s.get(key, default)

    gearshift = get("Gearshift") or get("Transmission Type", "")
    gears = get("Gear Ratios (Quantity)", "")
    transmission = f"{gears}-speed {gearshift}" if gears and gearshift else gearshift

    warranty_dist = get("Warranty Distance", "")
    warranty_years = get("Warranty Time (Years)", "")
    warranty = ""
    if warranty_dist:
        if warranty_years:
            warranty = f"{warranty_years}yr/{warranty_dist}km"
        else:
            warranty = f"{warranty_dist}km"

    sp_dist = get("Service Plan Distance", "")
    sp_years = get("Service Plan Time (Years)", "")
    service_plan = ""
    if sp_dist:
        if sp_years:
            service_plan = f"{sp_years}yr/{sp_dist}km"
        else:
            service_plan = f"{sp_dist}km"

    driven = get("Driven Wheels", "")
    drivetrain = driven
    if "front" in driven.lower():
        drivetrain = "FWD"
    elif "rear" in driven.lower():
        drivetrain = "RWD"
    elif "4" in driven or "all" in driven.lower():
        drivetrain = "AWD"

    return {
        "brand": brand_key,
        "model": model_name,
        "variant": variant_data.get("variant", ""),
        "price_incl": get("Price", "").replace("R", "").replace(",", "").replace(" ", "").strip(),
        "source": "quickpic_autospec",
        "scraped_at": datetime.now().isoformat(),
        "specifications": {
            "EngineCapacity": get("Engine Capacity (Cc)", ""),
            "PowerKW": get("Power Maximum", ""),
            "TorqueNM": get("Torque Maximum", ""),
            "FuelType": get("Fuel Type", ""),
            "Cylinders": get("Cylinders", ""),
            "Transmission": transmission,
            "Drivetrain": drivetrain,
            "TopSpeedKMH": get("Maximum/Top Speed", ""),
            "Acceleration0to100": get("Acceleration 0-100Km/h", ""),
            "FuelConsumptionL100KM": get("Fuel Consumption: Average", ""),
            "CO2EmissionsGKM": get("CO2 Emissions: Average", ""),
            "LengthMM": get("Length", ""),
            "WidthMM": get("Width Excl Mirrors / Incl Mirrors", ""),
            "HeightMM": get("Height", ""),
            "WheelbaseMM": get("Wheelbase", ""),
            "BootCapacityL": get("Load Volume Capacity", ""),
            "KerbWeightKG": get("Curb Weight", ""),
            "FuelTankL": get("Fuel Tank Capacity (Incl Reserve)", ""),
            "GroundClearanceMM": get("Ground Clearance Min-Max", ""),
            "TurningCircleM": get("Turning Circle (Wheels - Body)", ""),
            "Airbags": get("Airbag Quantity", ""),
            "ABS": get("Anti-Lock Braking System (ABS)", ""),
            "StabilityControl": get("Stability Control", ""),
            "Warranty": warranty,
            "ServicePlan": service_plan,
        }
    }


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f:
            return json.load(f)
    return {"completed": []}


def save_progress(progress):
    with open(PROGRESS_FILE, "w") as f:
        json.dump(progress, f, indent=2)


def build_smart_batches(work):
    """Group work items into batches of up to 4, preferring same-brand same-model grouping.
    This minimizes tree navigation clicks."""
    # Group by brand+model
    by_brand_model = defaultdict(list)
    for item in work:
        brand_key, qp_brand, model_name, variant = item
        by_brand_model[(qp_brand, model_name)].append(item)

    batches = []
    current_batch = []

    # Process each brand+model group
    groups = list(by_brand_model.values())
    random.shuffle(groups)

    for group in groups:
        for item in group:
            current_batch.append(item)
            if len(current_batch) == 4:
                batches.append(current_batch)
                current_batch = []

    if current_batch:
        batches.append(current_batch)

    return batches


async def expand_brand_in_dialog(page, qp_brand, model_name=""):
    """Search for a brand, expand it, then expand the specific model to show variants.
    Tree is 3 levels: Brand (ParentItem) > Model (ChildItem) > Variant (BabyItem).
    Uses keyboard.type() which properly triggers Blazor filtering.
    Returns number of variant checkboxes found (0 = failed)."""

    search_input = page.locator("input.main-input").first
    try:
        await search_input.wait_for(state="visible", timeout=8000)
    except Exception:
        print(f"    No search input found")
        return 0

    # Clear search box
    try:
        await search_input.click()
        await asyncio.sleep(random.uniform(0.5, 1.2))
        await search_input.click(click_count=3)
        await asyncio.sleep(random.uniform(0.3, 0.7))
        await page.keyboard.press("Backspace")
        await asyncio.sleep(random.uniform(2, 4))
    except Exception as e:
        print(f"    Clear error: {e}")
        return 0

    # Type brand name using keyboard.type() — this triggers Blazor filtering
    try:
        await search_input.click()
        await asyncio.sleep(random.uniform(0.3, 0.8))
        await page.keyboard.type(qp_brand, delay=random.randint(60, 140))
        await asyncio.sleep(random.uniform(4, 7))
    except Exception as e:
        print(f"    Type error: {e}")
        return 0

    # Check current state with detailed debug info
    state = await page.evaluate("""(brandName) => {
        const m = Array.from(document.querySelectorAll('.w3-modal'))
            .find(m => m.innerText.includes('Compare Vehicles'));
        if (!m) return {error: 'no modal', brands: 0, babies: 0, children: 0};

        const parents = m.querySelectorAll('.ParentItem');
        const brandTexts = Array.from(parents).map(p => {
            const n = p.querySelector('.chkboxItem');
            return n ? n.innerText.trim().substring(0, 40) : '?';
        });

        const children = m.querySelectorAll('.ChildItem');
        const childTexts = Array.from(children).map(c => {
            const n = c.querySelector('.chkboxItemChild') || c.querySelector('.chkboxItem');
            return n ? n.innerText.trim().substring(0, 40) : '?';
        });
        const babies = m.querySelectorAll('.BabyItem');
        const babyCheckboxes = m.querySelectorAll('.chkboxItemBaby');
        let cbCount = 0;
        for (const b of babyCheckboxes) {
            if (b.querySelector("input[type='checkbox']")) cbCount++;
        }

        return {
            brands: parents.length,
            brandTexts: brandTexts.slice(0, 5),
            children: children.length,
            childTexts: childTexts.slice(0, 8),
            babyItems: babies.length,
            babyCheckboxes: cbCount,
        };
    }""", qp_brand)

    baby_count = state.get("babyCheckboxes", 0)
    print(f"    Tree state: brands={state.get('brands')} children={state.get('children')} babies={state.get('babyItems')} checkboxes={baby_count}")
    if state.get("brandTexts"):
        print(f"    Brands: {state['brandTexts']}")

    # If babies already visible, check if they're from the right model
    if baby_count > 0:
        if baby_count > 100 or not model_name:
            return baby_count  # Full tree or no specific model needed
        # Small number of babies — might be wrong model. Check child texts for our model.
        child_texts = state.get("childTexts", [])
        # If we can see the model we need in the child list, the babies might be from another model
        model_lower = model_name.lower()
        # No way to check baby samples here, so fall through to expand the right model
        print(f"    Babies visible ({baby_count}) but need model {model_name}, checking...")

    # Check if models (ChildItems) are already visible — brand may have auto-expanded from search
    children_count = state.get("children", 0)
    if children_count > 0:
        print(f"    Brand already expanded with {children_count} models, skipping brand click")
    else:
        # No models visible — click brand to expand
        clicked = await page.evaluate("""(brandName) => {
            const m = Array.from(document.querySelectorAll('.w3-modal'))
                .find(m => m.innerText.includes('Compare Vehicles'));
            if (!m) return {clicked: false, error: 'no modal'};
            // Look in ParentItem for the brand
            const parents = m.querySelectorAll('.ParentItem');
            for (const parent of parents) {
                const item = parent.querySelector('.chkboxItem');
                if (item && item.innerText.includes(brandName)) {
                    // Click the expand-button on the brand, not the name (to avoid toggle issues)
                    const expandBtn = parent.querySelector('.expand-button');
                    if (expandBtn) {
                        expandBtn.click();
                        return {clicked: true, text: item.innerText.trim().substring(0, 40), method: 'expand-button'};
                    }
                    item.click();
                    return {clicked: true, text: item.innerText.trim().substring(0, 40), method: 'name-click'};
                }
            }
            return {clicked: false, error: 'brand not in tree'};
        }""", qp_brand)

        print(f"    Brand click: {clicked}")
        if not clicked or not clicked.get("clicked"):
            return 0

        # Wait for Blazor to render model list
        await asyncio.sleep(random.uniform(3, 6))

    # Check if models appeared but no babies yet — need to click the model too
    state2 = await page.evaluate("""() => {
        const m = Array.from(document.querySelectorAll('.w3-modal'))
            .find(m => m.innerText.includes('Compare Vehicles'));
        if (!m) return {babies: 0, children: 0, childTexts: []};
        const babies = m.querySelectorAll('.chkboxItemBaby');
        let count = 0;
        const samples = [];
        for (const b of babies) {
            if (b.querySelector("input[type='checkbox']")) {
                count++;
                if (samples.length < 3) {
                    const t = Array.from(b.childNodes).filter(n => n.nodeType === 3)
                        .map(n => n.textContent.trim()).filter(t => t).join(' ');
                    samples.push(t.substring(0, 50));
                }
            }
        }
        const children = m.querySelectorAll('.ChildItem');
        const childTexts = Array.from(children).map(c => {
            const n = c.querySelector('.chkboxItemChild') || c.querySelector('.chkboxItem');
            return n ? n.innerText.trim().substring(0, 40) : '?';
        });
        return {babies: count, children: children.length, childTexts: childTexts.slice(0, 10), samples: samples};
    }""")

    baby_count = state2.get("babies", 0)
    children_count = state2.get("children", 0)
    print(f"    After brand click: {baby_count} babies, {children_count} models, samples={state2.get('samples', [])}")
    if state2.get("childTexts"):
        print(f"    Models visible: {state2['childTexts'][:6]}")

    if baby_count > 0:
        # If full tree loaded (100+ babies) or no specific model needed, all variants visible
        if baby_count > 100 or not model_name:
            return baby_count
        # Small number of babies — might be from the WRONG model (e.g. Vitz instead of Hilux)
        # Check if any visible baby matches the target model name
        samples = state2.get("samples", [])
        model_lower = model_name.lower()
        if any(model_lower in s.lower() for s in samples):
            return baby_count  # Right model already expanded
        # Wrong model's babies visible — need to expand the correct model
        print(f"    Wrong model expanded ({baby_count} babies from {samples[:2]}), need {model_name}")

    # Models visible but wrong/no babies — click the specific MODEL's expand-button/dropdown
    # Model items use class .chkboxItemChild (NOT .chkboxItem) and have an .expand-button
    if children_count > 0 and model_name:
        print(f"    Clicking model dropdown: {model_name}")
        model_clicked = await page.evaluate("""(modelName) => {
            const m = Array.from(document.querySelectorAll('.w3-modal'))
                .find(m => m.innerText.includes('Compare Vehicles'));
            if (!m) return {clicked: false, error: 'no modal'};
            const children = m.querySelectorAll('.ChildItem');
            const normModel = modelName.toLowerCase().trim();

            // Try exact match first, then fuzzy
            for (const pass of ['exact', 'fuzzy']) {
                for (const child of children) {
                    // Model name is in .chkboxItemChild (not .chkboxItem)
                    const item = child.querySelector('.chkboxItemChild') || child.querySelector('.chkboxItem');
                    if (!item) continue;
                    const text = item.innerText.toLowerCase().trim();
                    const textBase = text.split('(')[0].trim();

                    let match = false;
                    if (pass === 'exact') {
                        match = textBase === normModel || text.startsWith(normModel + ' ') || text.startsWith(normModel + '(');
                    } else {
                        match = text.includes(normModel) || normModel.includes(textBase);
                    }

                    if (match) {
                        // Click the expand-button (dropdown arrow) to expand variants
                        const expandBtn = child.querySelector('.expand-button');
                        if (expandBtn) {
                            expandBtn.click();
                            return {clicked: true, text: item.innerText.trim().substring(0, 40), method: 'expand-button'};
                        }
                        // Fallback: click the model name itself
                        item.click();
                        return {clicked: true, text: item.innerText.trim().substring(0, 40), method: 'name-click'};
                    }
                }
            }
            return {clicked: false, error: 'model not found', available: Array.from(children).slice(0, 8).map(c => {
                const n = c.querySelector('.chkboxItemChild') || c.querySelector('.chkboxItem');
                return n ? n.innerText.trim().substring(0, 40) : '?';
            })};
        }""", model_name)

        print(f"    Model click: {model_clicked}")
        if model_clicked and model_clicked.get("clicked"):
            await asyncio.sleep(random.uniform(3, 6))

            # Recount babies after model click
            baby_count = await page.evaluate("""() => {
                const m = Array.from(document.querySelectorAll('.w3-modal'))
                    .find(m => m.innerText.includes('Compare Vehicles'));
                if (!m) return 0;
                let count = 0;
                for (const b of m.querySelectorAll('.chkboxItemBaby')) {
                    if (b.querySelector("input[type='checkbox']")) count++;
                }
                return count;
            }""")
            print(f"    After model click: {baby_count} babies")

            if baby_count > 0:
                return baby_count

    # Still no babies — try expand-button on brand as last resort
    if baby_count == 0:
        await page.evaluate("""(brandName) => {
            const m = Array.from(document.querySelectorAll('.w3-modal'))
                .find(m => m.innerText.includes('Compare Vehicles'));
            if (!m) return;
            const parents = m.querySelectorAll('.ParentItem');
            for (const parent of parents) {
                const nameEl = parent.querySelector('.chkboxItem');
                if (nameEl && nameEl.innerText.includes(brandName)) {
                    const btn = parent.querySelector('.expand-button');
                    if (btn) { btn.click(); return; }
                }
            }
        }""", qp_brand)
        await asyncio.sleep(random.uniform(3, 6))

        baby_count = await page.evaluate("""() => {
            const m = Array.from(document.querySelectorAll('.w3-modal'))
                .find(m => m.innerText.includes('Compare Vehicles'));
            if (!m) return 0;
            let count = 0;
            for (const b of m.querySelectorAll('.chkboxItemBaby')) {
                if (b.querySelector("input[type='checkbox']")) count++;
            }
            return count;
        }""")
        print(f"    After expand-button: {baby_count} babies")

    if baby_count == 0:
        # Save debug screenshot
        await page.screenshot(path=os.path.join(BASE_DIR, "debug_expand_fail.png"))

    return baby_count


async def select_variant_checkbox(page, variant_name, model_name=""):
    """Find and check a variant's checkbox in the expanded tree.
    Variants are .BabyItem > .chkboxItemBaby elements.
    Scopes search to the correct model's section if model_name provided.
    Returns dict with found status and matched text."""

    result = await page.evaluate("""(args) => {
        const variantName = args.variant;
        const modelName = args.model;
        const m = Array.from(document.querySelectorAll('.w3-modal'))
            .find(m => m.innerText.includes('Compare Vehicles'));
        if (!m) return {found: false, error: 'no modal'};

        // Get all items in the tree (they're flat siblings in .list-con)
        const listCon = m.querySelector('.list-con');
        if (!listCon) return {found: false, error: 'no list-con'};

        // Build model-scoped baby list: find the ChildItem matching our model,
        // then collect BabyItems after it until next ChildItem/ParentItem
        let scopedBabies = [];
        let allBabies = [];
        const children = listCon.children;

        if (modelName) {
            let inModel = false;
            const normModel = modelName.toLowerCase().trim();
            for (const el of children) {
                const cls = el.className || '';
                if (cls.includes('ChildItem') || cls.includes('ParentItem')) {
                    if (inModel) break;  // Left our model section
                    if (cls.includes('ChildItem')) {
                        const modelText = el.innerText.toLowerCase().trim();
                        if (modelText.startsWith(normModel) || modelText.includes(normModel + ' ')) {
                            inModel = true;
                        }
                    }
                } else if (cls.includes('BabyItem') && inModel) {
                    const baby = el.querySelector('.chkboxItemBaby');
                    if (baby) scopedBabies.push(baby);
                }
            }
        }

        // Also collect ALL babies as fallback
        const allBabyEls = m.querySelectorAll('.chkboxItemBaby');
        for (const b of allBabyEls) {
            if (b.querySelector("input[type='checkbox']")) allBabies.push(b);
        }

        // Search scoped first, then all as fallback
        const searchSets = scopedBabies.length > 0
            ? [{babies: scopedBabies, bonus: 10}, {babies: allBabies, bonus: 0}]
            : [{babies: allBabies, bonus: 0}];

        let bestMatch = null;
        let bestScore = 0;

        for (const {babies, bonus} of searchSets) {
            for (const baby of babies) {
                const cb = baby.querySelector("input[type='checkbox']");
                if (!cb || cb.checked) continue;

                const text = Array.from(baby.childNodes)
                    .filter(n => n.nodeType === 3)
                    .map(n => n.textContent.trim())
                    .filter(t => t)
                    .join(' ').trim();
                if (!text) continue;

                let score = 0;
                const normText = text.toLowerCase().replace(/\\s+/g, ' ').trim();
                const normVar = variantName.toLowerCase().replace(/\\s+/g, ' ').trim();

                if (normText === normVar) score = 100;
                else if (normText.includes(normVar) && normVar.length >= 5) score = 80;
                else if (normVar.includes(normText) && normText.length >= 10) score = 70;
                else {
                    const textWords = normText.split(/\\s+/);
                    const varWords = normVar.split(/\\s+/);
                    const matches = varWords.filter(w => textWords.some(tw => tw === w) && w.length > 1);
                    if (matches.length >= 3 || (matches.length >= 2 && matches.length >= varWords.length * 0.5)) {
                        score = 30 + matches.length * 5;
                    }
                }

                score += bonus;  // Boost for model-scoped matches

                if (score > bestScore) {
                    bestScore = score;
                    bestMatch = {cb, text, score};
                }
            }

            // If found a good match in scoped set, use it
            if (bestMatch && bestScore >= 40) break;
        }

        if (bestMatch && bestScore >= 40) {
            bestMatch.cb.click();
            return {found: true, text: bestMatch.text.substring(0, 100), score: bestScore};
        }

        // Debug info
        const sampleBabies = [];
        for (const baby of (scopedBabies.length > 0 ? scopedBabies : allBabies).slice(0, 8)) {
            const t = Array.from(baby.childNodes)
                .filter(n => n.nodeType === 3).map(n => n.textContent.trim())
                .filter(t => t).join(' ').trim();
            if (t) sampleBabies.push(t.substring(0, 60));
        }
        return {found: false, total: allBabies.length, scoped: scopedBabies.length,
                variants: sampleBabies};
    }""", {"variant": variant_name, "model": model_name})

    return result


async def browse_around(page):
    """Browse away from AutoSpec like a real human — visit homepage, read an article,
    then come back. Resets Blazor state naturally."""
    print("    [BROWSE] Taking a break from AutoSpec...")

    try:
        # Go to homepage / news
        await page.goto("https://www.quickpic.co.za/news", wait_until="domcontentloaded", timeout=90000)
        await asyncio.sleep(random.uniform(4, 8))

        # Try to click into a random article
        try:
            articles = await page.locator("a[href*='/news/']").all()
            if articles:
                pick = random.choice(articles[:12])  # Pick from top articles
                await pick.scroll_into_view_if_needed()
                await asyncio.sleep(random.uniform(1, 3))
                await pick.click()
                # "Read" the article for a bit
                read_time = random.uniform(25, 65)
                print(f"    [BROWSE] Reading article for {read_time:.0f}s...")
                await asyncio.sleep(read_time)
                # Scroll down a bit like actually reading
                for _ in range(random.randint(2, 5)):
                    await page.mouse.wheel(0, random.randint(200, 500))
                    await asyncio.sleep(random.uniform(2, 6))
            else:
                # Just hang on the news page
                await asyncio.sleep(random.uniform(20, 45))
        except Exception as e:
            print(f"    [BROWSE] Article click failed ({e}), just waiting...")
            await asyncio.sleep(random.uniform(15, 30))

        # Now go back to AutoSpec fresh
        print("    [BROWSE] Back to AutoSpec...")
        await page.goto("https://www.quickpic.co.za/autospec", wait_until="domcontentloaded", timeout=90000)
        await asyncio.sleep(random.uniform(5, 10))
    except Exception as e:
        # Network error during browse — wait and try to get back to AutoSpec
        print(f"    [BROWSE] Navigation failed ({type(e).__name__}), waiting and retrying...")
        await asyncio.sleep(random.uniform(30, 60))
        try:
            await page.goto("https://www.quickpic.co.za/autospec", wait_until="domcontentloaded", timeout=120000)
            await asyncio.sleep(random.uniform(5, 10))
        except Exception:
            print("    [BROWSE] Still can't reach QuickPic, will retry on next batch")


async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--brand", help="Process single brand key")
    parser.add_argument("--limit", type=int, default=0, help="Max variants to fetch")
    args = parser.parse_args()

    with open(CATALOGUE_FILE) as f:
        catalogue = json.load(f)

    progress = load_progress()
    completed_set = set(progress["completed"])

    work = []
    for brand_key, brand_data in catalogue.items():
        if brand_key.startswith("_extra_"):
            continue
        if args.brand and brand_key != args.brand:
            continue

        for model_name, model_data in brand_data.get("models", {}).items():
            for variant in model_data.get("variants", []):
                vid = f"{brand_key}|{model_name}|{variant['name']}"
                if vid not in completed_set:
                    work.append((brand_key, brand_data["qp_name"], model_name, variant))

    if args.limit:
        random.shuffle(work)
        work = work[:args.limit]

    print(f"Work queue: {len(work)} variants to fetch")
    if not work:
        print("Nothing to do!")
        return

    batches = build_smart_batches(work)
    print(f"Batches: {len(batches)} (up to 4 vehicles each)")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            viewport={"width": 1366, "height": 768},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            locale="en-ZA",
            timezone_id="Africa/Johannesburg",
            accept_downloads=True,
        )
        page = await context.new_page()

        # Login — slow, human-like
        print("\n[LOGIN]")
        await page.goto("https://www.quickpic.co.za", wait_until="domcontentloaded", timeout=90000)
        await asyncio.sleep(random.uniform(6, 12))
        await page.locator("button:has-text('LOGIN')").first.click()
        await asyncio.sleep(random.uniform(3, 6))
        await human_type(page.locator("input[name='email']").first, USERNAME)
        await asyncio.sleep(random.uniform(2, 4))
        await human_type(page.locator("input[type='password']").first, PASSWORD)
        await asyncio.sleep(random.uniform(2, 5))
        await page.locator("button:has-text('Login')").last.click()
        await asyncio.sleep(random.uniform(12, 20))
        print(f"  Logged in: {page.url}")

        total_fetched = 0
        consecutive_fails = 0
        batches_since_browse = 0
        next_browse_at = random.randint(3, 5)  # Browse after 3-5 batches

        # Initial navigate to AutoSpec
        await page.goto("https://www.quickpic.co.za/autospec", wait_until="domcontentloaded", timeout=90000)
        await asyncio.sleep(random.uniform(5, 10))

        for batch_idx, batch in enumerate(batches):
            print(f"\n{'='*50}")
            print(f"  Batch {batch_idx+1}/{len(batches)} — {len(batch)} vehicles")
            for _, qp_brand, model, var in batch:
                print(f"    {qp_brand} {model} {var['name']}")
            print(f"{'='*50}")

            # Human-like delay between batches
            await wild_delay("between batches")

            # Periodically browse away to look human and reset Blazor state
            batches_since_browse += 1
            if batches_since_browse >= next_browse_at:
                await browse_around(page)
                batches_since_browse = 0
                next_browse_at = random.randint(3, 5)
            else:
                # Just clear comparison by navigating within AutoSpec
                await page.goto("https://www.quickpic.co.za/autospec", wait_until="domcontentloaded", timeout=90000)
                await asyncio.sleep(random.uniform(3, 6))

            selected = 0
            selected_items = []

            # Each checkbox click CLOSES the dialog and adds vehicle to comparison.
            # Flow: open dialog → search brand → click variant checkbox → dialog closes
            # Repeat for each variant (up to 4 per batch).

            for item_idx, (brand_key, qp_brand, model_name, variant) in enumerate(batch):
                variant_name = variant["name"]

                if item_idx > 0:
                    # Human-like delay between vehicle selections
                    await asyncio.sleep(random.uniform(3, 7))

                # Open compare dialog
                print(f"    [{item_idx+1}/{len(batch)}] Opening dialog...")
                try:
                    await page.locator("#SelectVehicleLabel").first.click(timeout=10000)
                except Exception:
                    try:
                        await page.evaluate("document.getElementById('SelectVehicleLabel')?.click()")
                    except Exception:
                        print(f"    Can't open dialog for item {item_idx+1}")
                        continue
                await asyncio.sleep(random.uniform(3, 5))

                # Expand the brand + model tree
                baby_count = await expand_brand_in_dialog(page, qp_brand, model_name)
                if baby_count == 0:
                    print(f"    No variants found, refreshing page and retrying...")
                    # Full page refresh to reset Blazor state
                    await page.goto("https://www.quickpic.co.za/autospec", wait_until="domcontentloaded", timeout=90000)
                    await asyncio.sleep(random.uniform(5, 10))
                    # Re-select any previously selected vehicles in this batch
                    for prev_brand_key, prev_qp_brand, prev_model, prev_var in selected_items:
                        try:
                            await page.locator("#SelectVehicleLabel").first.click(timeout=10000)
                            await asyncio.sleep(random.uniform(3, 5))
                            await expand_brand_in_dialog(page, prev_qp_brand, prev_model)
                            await select_variant_checkbox(page, prev_var["name"], prev_model)
                            await asyncio.sleep(random.uniform(3, 6))
                        except Exception:
                            pass
                    # Now retry current vehicle
                    try:
                        await page.locator("#SelectVehicleLabel").first.click(timeout=10000)
                        await asyncio.sleep(random.uniform(3, 5))
                    except Exception:
                        print(f"    SKIP: Can't reopen dialog for {qp_brand}")
                        continue
                    baby_count = await expand_brand_in_dialog(page, qp_brand, model_name)
                    if baby_count == 0:
                        print(f"    SKIP: Can't expand {qp_brand} {model_name} even after refresh")
                        continue

                # Select the variant checkbox (this will close the dialog)
                print(f"    Selecting: {model_name} {variant_name}")
                result = await select_variant_checkbox(page, variant_name, model_name)

                if result and result.get("found"):
                    selected += 1
                    selected_items.append((brand_key, qp_brand, model_name, variant))
                    print(f"    OK: {result.get('text', variant_name)[:60]} (score={result.get('score')})")
                    # Wait for dialog to close and page to update
                    await asyncio.sleep(random.uniform(4, 8))
                else:
                    total_babies = result.get("total", 0) if result else 0
                    variants_vis = result.get("variants", []) if result else []
                    print(f"    MISS: {variant_name[:50]} ({total_babies} variants in tree)")
                    if variants_vis:
                        print(f"    Sample: {variants_vis[:3]}")

            if selected == 0:
                print("  No vehicles selected, skipping batch")
                consecutive_fails += 1
                if consecutive_fails >= 2:
                    # Browse around to reset state before giving up
                    print("  2 consecutive fails — browsing around to reset...")
                    await browse_around(page)
                    batches_since_browse = 0
                if consecutive_fails >= 5:
                    print("  5 consecutive failures, stopping")
                    break
                continue

            consecutive_fails = 0

            # Comparison table should already be showing. Expand all spec sections.
            print(f"  {selected} vehicles on page, expanding sections...")
            await asyncio.sleep(random.uniform(3, 6))

            sections = ["General", "Engine", "Safety", "Warranty", "Features",
                        "Aesthetics And Comfort", "Specifications", "Wheels And Tyres"]
            for sec in sections:
                try:
                    sec_el = page.locator(f"text='{sec}'").first
                    if await sec_el.count() > 0:
                        await sec_el.click()
                        await asyncio.sleep(random.uniform(0.5, 1.5))
                except Exception:
                    pass

            await medium_delay()

            # Screenshot for debugging the comparison table
            ss_path = os.path.join(BASE_DIR, f"debug_compare_b{batch_idx+1}.png")
            await page.screenshot(path=ss_path, full_page=True)
            print(f"  Screenshot: {ss_path}")

            # Scrape HTML comparison table
            print("  Scraping comparison table...")
            try:
                # Dump raw table HTML for debugging (shows innerHTML to reveal ranking elements)
                table_html = await page.evaluate("""() => {
                    const tables = document.querySelectorAll('table');
                    let biggest = null;
                    for (const t of tables) {
                        if (!biggest || t.innerHTML.length > biggest.innerHTML.length) biggest = t;
                    }
                    if (!biggest) return '';
                    const rows = biggest.querySelectorAll('tr');
                    let html = '';
                    for (let i = 0; i < Math.min(rows.length, 25); i++) {
                        const cells = rows[i].querySelectorAll('td, th');
                        const texts = Array.from(cells).map(c => c.innerText.trim().substring(0, 60));
                        const htmls = Array.from(cells).map(c => c.innerHTML.substring(0, 120));
                        html += `Row ${i} text: ${JSON.stringify(texts)}\\n`;
                        html += `Row ${i} html: ${JSON.stringify(htmls)}\\n`;
                    }
                    return html;
                }""")
                if table_html:
                    debug_path = os.path.join(BASE_DIR, "debug_table_rows.txt")
                    with open(debug_path, "w", encoding="utf-8") as f:
                        f.write(table_html)
                    print(f"  Table rows saved to: {debug_path}")

                vehicles = await scrape_comparison_html(page, selected_items)

                # If no specs, page may not have loaded — wait and retry
                if vehicles and all(len(v["specs"]) == 0 for v in vehicles):
                    print("  Table empty, waiting for data to load...")
                    await asyncio.sleep(random.uniform(8, 14))
                    vehicles = await scrape_comparison_html(page, selected_items)

                print(f"  Parsed: {len(vehicles)} vehicles from HTML")

                for vdata in vehicles:
                    # vdata["variant"] is the variant name from selected_items
                    variant_name = vdata["variant"]
                    # Find the matching batch item
                    for brand_key, qp_brand, model_name, variant in selected_items:
                        if variant["name"] == variant_name:
                            spec_data = specs_to_feed_format(brand_key, model_name, vdata)
                            safe_name = f"{brand_key}_{model_name}_{variant['name']}"
                            safe_name = re.sub(r'[^\w\-.]', '_', safe_name)
                            spec_file = os.path.join(SPECS_DIR, f"{safe_name}.json")
                            with open(spec_file, "w") as f:
                                json.dump(spec_data, f, indent=2)

                            merged = merge_specs_to_raw(brand_key, model_name,
                                                        variant["name"], vdata["specs"])

                            vid = f"{brand_key}|{model_name}|{variant['name']}"
                            progress["completed"].append(vid)
                            total_fetched += 1
                            m_flag = " [merged]" if merged else ""
                            print(f"    Saved: {brand_key} {model_name} {variant['name']}{m_flag}")
                            break

                save_progress(progress)

            except Exception as e:
                print(f"  HTML scrape failed: {e}")
                import traceback
                traceback.print_exc()

            print(f"  Total fetched so far: {total_fetched}")

        await browser.close()

    print(f"\n{'='*60}")
    print(f"  DONE — Fetched specs for {total_fetched} variants")
    print(f"  Specs saved to: {SPECS_DIR}")
    print(f"{'='*60}")


if __name__ == "__main__":
    asyncio.run(main())
